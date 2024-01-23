from openpyxl import Workbook
import datetime
import random
import os

workbook = Workbook()

sheet = workbook.active
sheet.title = "TDSheet"

names = ['Olzhas', 'Asd', 'Nur', 'AAA', 'Adilet', 'Nural', 'Arna', 'Po', 'ASD', 'Fff']

random_name = random.choice(names)

sheet.cell(row=1, column=1).value = random_name
sheet.cell(row=1, column=2).value = datetime.datetime.now().strftime('%Y-%m-%d')
sheet.cell(row=1, column=3).value = datetime.datetime.now().strftime('%H:%M:%S')

file_name = f"{random_name}_{datetime.datetime.now().strftime('%Y%m%d')}_{random.randint(100, 999)}.xlsx"
save_path = os.path.join(os.path.expanduser('~'), 'Documents', 'skcu', file_name)
os.makedirs(os.path.dirname(save_path), exist_ok=True)
workbook.save(save_path)
print(f"Файл {file_name} создан и сохранен в {save_path}")


